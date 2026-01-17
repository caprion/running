#!/usr/bin/env python3
"""Debug script to inspect Excel structure"""

import openpyxl
from pathlib import Path

excel_path = Path("resources/20_Week_Training_Plan.xlsx")
workbook = openpyxl.load_workbook(excel_path, data_only=True)

print("Available sheets:")
for sheet_name in workbook.sheetnames:
    print(f"  - {sheet_name}")
print()

# Look at each sheet
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    print(f"\nSheet: {sheet_name}")
    print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")
    print("=" * 120)

    # Print first 10 rows
    for row_idx in range(1, min(11, sheet.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(10, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = str(cell.value)[:15] if cell.value else ""
            row_values.append(value)

        print(f"Row {row_idx:2d}: {' | '.join(row_values)}")

workbook.close()
